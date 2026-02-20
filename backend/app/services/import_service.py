"""Import service: parse Excel/CSV files into shift data."""

import io
from datetime import datetime
from typing import List, Dict, Any, Tuple


def parse_excel(file_bytes: bytes, name_filter: str = None) -> Tuple[List[Dict], List[str]]:
    """Parse an .xlsx file and return (shifts, errors)."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], ["Filen er tom"]

    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    return _parse_rows(header, rows[1:], name_filter)


def parse_csv(file_bytes: bytes, name_filter: str = None) -> Tuple[List[Dict], List[str]]:
    """Parse a CSV file and return (shifts, errors)."""
    import csv

    text = file_bytes.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text), delimiter=";")
    rows = list(reader)
    if not rows:
        return [], ["Filen er tom"]

    header = [c.strip().lower() for c in rows[0]]
    return _parse_rows(header, rows[1:], name_filter)


def _find_col(header: List[str], candidates: List[str]) -> int:
    for c in candidates:
        if c in header:
            return header.index(c)
    return -1


def _parse_rows(header: List[str], rows: list, name_filter: str) -> Tuple[List[Dict], List[str]]:
    date_col = _find_col(header, ["dato", "date"])
    start_col = _find_col(header, ["start", "starttid", "fra", "from"])
    end_col = _find_col(header, ["slutt", "end", "til", "to", "stopp"])
    pause_col = _find_col(header, ["pause", "pause_min", "break"])
    name_col = _find_col(header, ["navn", "name", "ansatt"])
    note_col = _find_col(header, ["notat", "note", "kommentar"])

    shifts = []
    errors = []

    for i, row in enumerate(rows, 2):
        if not any(row):
            continue

        def get(col):
            if col < 0 or col >= len(row):
                return None
            return row[col]

        # Name filter
        if name_filter and name_col >= 0:
            name_val = str(get(name_col) or "").strip()
            if name_val and name_filter.lower() not in name_val.lower():
                continue

        # Date
        raw_date = get(date_col)
        if raw_date is None:
            errors.append(f"Rad {i}: mangler dato")
            continue
        date_str = _parse_date(raw_date)
        if not date_str:
            errors.append(f"Rad {i}: ugyldig datoformat ({raw_date})")
            continue

        # Start / end
        raw_start = get(start_col)
        raw_end = get(end_col)
        if raw_start is None or raw_end is None:
            errors.append(f"Rad {i}: mangler start/slutt-tid")
            continue
        start_str = _parse_time(raw_start)
        end_str = _parse_time(raw_end)
        if not start_str or not end_str:
            errors.append(f"Rad {i}: ugyldig tidsformat ({raw_start} / {raw_end})")
            continue

        # Pause
        pause_min = 0
        if pause_col >= 0:
            try:
                pause_min = int(float(get(pause_col) or 0))
            except (ValueError, TypeError):
                pass

        shifts.append({
            "date": date_str,
            "start_time": start_str,
            "end_time": end_str,
            "pause_min": pause_min,
            "note": str(get(note_col) or "").strip() or None,
        })

    return shifts, errors


def _parse_date(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return ""


def _parse_time(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    s = str(value).strip()
    for fmt in ("%H:%M", "%H.%M", "%H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).strftime("%H:%M")
        except ValueError:
            pass
    return ""
