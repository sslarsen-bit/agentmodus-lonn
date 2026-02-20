"""Export service: generate PDF, CSV and Excel reports."""

import io
import csv
from datetime import datetime
from typing import List

from ..models.shift import Shift
from ..models.user import User
from ..models.month_summary import MonthSummary


def generate_csv(shifts: List[Shift], user: User) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow([
        "Dato", "Start", "Slutt", "Pause (min)",
        "Timer totalt", "Grunntimer", "Kveldstimer", "Nattimer",
        "Helgetimer", "Helligdagstimer", "OT50 timer", "OT100 timer",
        "Brutto (kr)", "Notat"
    ])
    for s in shifts:
        writer.writerow([
            s.date, s.start_time, s.end_time, s.pause_min,
            f"{s.total_hours:.2f}", f"{s.base_hours:.2f}",
            f"{s.evening_hours:.2f}", f"{s.night_hours:.2f}",
            f"{s.weekend_hours:.2f}", f"{s.holiday_hours:.2f}",
            f"{s.overtime_50_hours:.2f}", f"{s.overtime_100_hours:.2f}",
            f"{s.gross_pay:.2f}", s.note or ""
        ])
    return output.getvalue().encode("utf-8-sig")


def generate_excel(shifts: List[Shift], user: User, summary: MonthSummary = None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Vakter"

    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = [
        "Dato", "Start", "Slutt", "Pause (min)",
        "Timer totalt", "Grunntimer", "Kveldstimer", "Nattimer",
        "Helgetimer", "Helligdagstimer", "OT50", "OT100",
        "Brutto (kr)", "Notat"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = 14

    for row_idx, s in enumerate(shifts, 2):
        row_data = [
            s.date, s.start_time, s.end_time, s.pause_min,
            round(s.total_hours, 2), round(s.base_hours, 2),
            round(s.evening_hours, 2), round(s.night_hours, 2),
            round(s.weekend_hours, 2), round(s.holiday_hours, 2),
            round(s.overtime_50_hours, 2), round(s.overtime_100_hours, 2),
            round(s.gross_pay, 2), s.note or ""
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = border

    if summary:
        ws2 = wb.create_sheet(title="Sammendrag")
        items = [
            ("Navn", user.name),
            ("Periode", f"{summary.year}-{summary.month:02d}"),
            ("Totale timer", f"{summary.total_hours:.2f}"),
            ("Overtid 50%", f"{summary.overtime_50_hours:.2f}"),
            ("Overtid 100%", f"{summary.overtime_100_hours:.2f}"),
            ("Bruttolønn", f"{summary.gross_pay:.2f} kr"),
            ("Skattetrekk (est.)", f"{summary.tax_deduction:.2f} kr"),
            ("Netto (est.)", f"{summary.net_pay:.2f} kr"),
            ("Feriepenger opptjent", f"{summary.holiday_pay_earned:.2f} kr"),
        ]
        for r, (label, val) in enumerate(items, 1):
            ws2.cell(row=r, column=1, value=label).font = Font(bold=True)
            ws2.cell(row=r, column=2, value=val)
        ws2.column_dimensions["A"].width = 24
        ws2.column_dimensions["B"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_pdf(shifts: List[Shift], user: User, summary: MonthSummary = None) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    elements = []

    title_style = styles["Heading1"]
    elements.append(Paragraph(f"Vaktrapport – {user.name}", title_style))
    if summary:
        elements.append(Paragraph(f"Periode: {summary.year}-{summary.month:02d}", styles["Normal"]))
    elements.append(Spacer(1, 0.5*cm))

    if summary:
        summary_data = [
            ["Totale timer", f"{summary.total_hours:.2f} t"],
            ["Overtid 50%", f"{summary.overtime_50_hours:.2f} t"],
            ["Overtid 100%", f"{summary.overtime_100_hours:.2f} t"],
            ["Bruttolønn", f"{summary.gross_pay:.2f} kr"],
            ["Skattetrekk (est.)", f"{summary.tax_deduction:.2f} kr"],
            ["Netto (est.)", f"{summary.net_pay:.2f} kr"],
            ["Feriepenger opptjent", f"{summary.holiday_pay_earned:.2f} kr"],
        ]
        t = Table(summary_data, colWidths=[8*cm, 6*cm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.whitesmoke),
            ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#4F46E5")),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#F5F5FF")]),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 0.5*cm))

    elements.append(Paragraph("Vaktdetaljer", styles["Heading2"]))
    elements.append(Spacer(1, 0.3*cm))

    data = [["Dato", "Start", "Slutt", "Timer", "Brutto (kr)", "Notat"]]
    for s in shifts:
        data.append([
            s.date, s.start_time, s.end_time,
            f"{s.total_hours:.2f}", f"{s.gross_pay:.2f}",
            (s.note or "")[:30]
        ])

    t2 = Table(data, repeatRows=1)
    t2.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5FF")]),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    elements.append(t2)

    elements.append(Spacer(1, 0.5*cm))
    elements.append(Paragraph("*Skattetrekk og netto er estimater og erstatter ikke offisielt lønnssystem.",
                               styles["Italic"]))

    doc.build(elements)
    return buf.getvalue()
